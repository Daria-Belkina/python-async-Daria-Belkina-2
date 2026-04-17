import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import asyncio
import time


class MedicalDevices:
    """Класс для обработки данных медицинских клиник и оборудования асинхронно."""

    def __init__(self, filepath: str = 'NA') -> None:
        """Конструктор класса.

        Args:
            filepath: Имя файла.
        """

        self.filepath = filepath
        self.df = None

    def read_file(self) -> pd.DataFrame:
        """Функция для чтения данных из Excel файла.

        Returns:
            self.df: Объект DataFrame с данными из файла.
        """

        self.df = pd.read_excel(self.filepath)

        return self.df

    async def read_one_file(self, filepath) -> pd.DataFrame:
        """Асинхронное чтение одного Excel файла.

        Args:
            filepath: Путь к файлу.

        Returns:
            DataFrame с данными из файла.
        """

        return await asyncio.to_thread(pd.read_excel, filepath)

    async def read_all_files(self, filepaths):
        """Асинхронное чтение всех Excel файлов.

        Args:
            filepaths: Список путей к файлам.

        Returns:
            Объединенный DataFrame из всех файлов.
        """

        tasks = [self.read_one_file(filepath) for filepath in filepaths]
        data = await asyncio.gather(*tasks)
        self.df = pd.concat(data, ignore_index=True)

        return self.df

    def status_normalization(self) -> pd.DataFrame:
        """Функция для нормализации статусов устройств.

        Returns:
            self.df: Объект DataFrame с нормализованными статусами.
        """

        status_map = {
            'planned_installation': ['planned_installation', 'planned', 'scheduled_install', 'to_install'],
            'operational': ['operational', 'op', 'working', 'OK'],
            'maintenance_scheduled': ['maintenance_scheduled', 'maint_sched', 'maintenance', 'service_scheduled'],
            'faulty': ['faulty', 'error', 'broken', 'needs_repair']
        }
        for normalized, variants in status_map.items():
            self.df.loc[self.df['status'].isin(variants), 'status'] = normalized

        return self.df

    def correcting_date_format(self, date_value) -> pd.DataFrame:
        """Функция для исправления формата даты.

        Args:
            date_value: Значение даты в любом формате.

        Returns:
            datetime.strptime: Дата в исправленном формате.
        """

        if pd.isna(date_value) or date_value == '' or date_value == 0:

            return np.nan

        date_str = str(date_value).strip()
        formats = ['%Y-%m-%d', '%d.%m.%Y', '%b %d, %Y', '%d %b %Y', '%Y/%m/%d', '%m/%d/%Y', '%d-%b-%Y', '%Y%m%d']

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        return np.nan

    def parse_dates(self) -> pd.DataFrame:
        """Отдельная функция для исправления всех дат.

        Returns:
            self.df: Объект DataFrame с нормализованными датами.
        """

        date_cols = ['install_date', 'warranty_until', 'last_calibration_date', 'last_service_date']
        for col in date_cols:
            if col in self.df.columns:
                self.df[col] = self.df[col].apply(self.correcting_date_format)
                self.df[col] = pd.to_datetime(self.df[col]).dt.date
                print(f"Столбец {col} обработан!")

        return self.df

    def checking_calibration_date(self) -> pd.DataFrame:
        """Функция для проверки корректности дат калибровки.

        Returns:
            self.df: Объект DataFrame с нормализованными датами калибровки.
        """

        invalid = (self.df['install_date'].notna() & self.df['last_calibration_date'].notna() &
                   (self.df['install_date'] > self.df['last_calibration_date']))
        self.df.loc[invalid, 'last_calibration_date'] = np.nan

        return self.df

    def checking_service_date(self) -> pd.DataFrame:
        """Функция для проверки корректности дат обслуживания.

        Returns:
            self.df: объект DataFrame с нормализованными датами обслуживания.
        """

        invalid = (self.df['install_date'].notna() & self.df['last_service_date'].notna() &
                   (self.df['install_date'] > self.df['last_service_date']))
        self.df.loc[invalid, 'last_service_date'] = np.nan

        return self.df

    def filter_under_warranty_sync(self) -> pd.DataFrame:
        """Функция для фильтрации устройств по гарантии.

        Returns:
            filepath_warranty: DataFrame с устройствами на гарантии.
        """

        today_date = datetime.today().date()

        return self.df[self.df['warranty_until'] >= today_date].copy()

    def problems_of_clinics_sync(self, df_to_analyze) -> pd.DataFrame:
        """Функция для анализа проблем по клиникам.

        Args:
            df_to_analyze (pandas.DataFrame): DataFrame для анализа.

        Returns:
            filepath_problems: DataFrame с клиниками и суммой проблем в них.
        """

        filepath_problems = df_to_analyze.groupby('clinic_id').agg({
            'issues_reported_12mo': 'sum', 'clinic_name': 'first', 'city': 'first'
        }).reset_index()
        filepath_problems.columns = ['clinic_id', 'cnt_problems', 'clinic_name', 'city']

        return filepath_problems.sort_values('cnt_problems', ascending=False)

    def calibration_dates_sync(self, df_to_analyze) -> pd.DataFrame:
        """Функция для анализа количества дней с последней калибровки.

        Args:
            df_to_analyze: DataFrame для анализа.

        Returns:
            filepath_calibration: Исходный DataFrame с дополнительной колонкой (количество дней с последней калибровки).
        """

        today_date = datetime.today().date()
        res = df_to_analyze.copy()
        res['days_since_last_calibration'] = today_date - res['last_calibration_date']

        return res

    def create_pivot_table_sync(self, df_to_pivot) -> pd.DataFrame:
        """Функция для создания сводной таблицы.

        Args:
            df_to_pivot: DataFrame для создания сводной таблицы.

        Returns:
            pivot_table: Сводная таблица.
        """

        return df_to_pivot.groupby(['clinic_id', 'clinic_name', 'model', 'device_id']).agg({
            'issues_reported_12mo': 'sum', 'failure_count_12mo': 'sum', 'uptime_pct': 'mean'
        }).reset_index()

    async def filter_under_warranty(self) -> pd.DataFrame:
        """Асинхронная фильтрация устройств по гарантии.

        Returns:
            DataFrame с устройствами, находящимися на гарантии.
        """

        return await asyncio.to_thread(self.filter_under_warranty_sync)

    async def problems_of_clinics(self, df_to_analyze) -> pd.DataFrame:
        """Асинхронный анализ проблем по клиникам.

        Args:
            df_to_analyze: DataFrame для анализа.

        Returns:
            DataFrame с клиниками и количеством проблем, отсортированный по убыванию.
        """

        return await asyncio.to_thread(self.problems_of_clinics_sync, df_to_analyze)

    async def calibration_dates(self, df_to_analyze) -> pd.DataFrame:
        """Асинхронный отчёт по срокам калибровки.

        Args:
            df_to_analyze: DataFrame для анализа.

        Returns:
            DataFrame с дополнительной колонкой дней с последней калибровки.
        """

        return await asyncio.to_thread(self.calibration_dates_sync, df_to_analyze)

    async def create_pivot_table(self, df_to_pivot) -> pd.DataFrame:
        """Асинхронное создание сводной таблицы.

        Args:
            df_to_pivot: DataFrame для создания сводной таблицы.

        Returns:
            Сводная таблица с агрегированными данными по клиникам и оборудованию.
        """

        return await asyncio.to_thread(self.create_pivot_table_sync, df_to_pivot)


def _format_excel_pivot(filepath: str):
    """Форматирование сводной таблицы в Excel.

    Объединяет ячейки с одинаковыми значениями в колонках A, B, C.

    Args:
        filepath: Путь к Excel файлу для форматирования.
    """

    wb = load_workbook(filepath)
    ws = wb.active
    for col_letter in ['A', 'B', 'C']:
        current_value = None
        start_row = 2
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'{col_letter}{row}'].value
            if cell_value != current_value:
                if start_row < row - 1:
                    ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{row - 1}')
                    ws[f'{col_letter}{start_row}'].alignment = Alignment(horizontal='center')
                current_value = cell_value
                start_row = row
    wb.save(filepath)

async def _save_and_format_pivot(df, filepath: str):
    """Асинхронное сохранение и форматирование сводной таблицы.

    Args:
        df: DataFrame для сохранения.
        filepath: Путь для сохранения файла.
    """

    await asyncio.to_thread(df.to_excel, filepath, index=False, engine='openpyxl')
    await asyncio.to_thread(_format_excel_pivot, filepath)


def run_sync(file_list):
    """Синхронная обработка всех файлов.

    Args:
        file_list: Список путей к Excel файлам.

    Returns:
        Время выполнения в секундах.
    """

    print("ЗАПУСК СИНХРОННОЙ ВЕРСИИ\n" + "=" * 60)
    start_time = time.time()

    print("\nЧтение файлов...")
    all_dfs = []
    for fp in file_list:
        process_sync = MedicalDevices(fp)
        process_sync.read_file()
        all_dfs.append(process_sync.df)
    combined_df = pd.concat(all_dfs, ignore_index=True)

    process_sync = MedicalDevices('combined')
    process_sync.df = combined_df
    process_sync.parse_dates()
    process_sync.status_normalization()
    process_sync.checking_calibration_date()
    process_sync.checking_service_date()
    process_sync.df.dropna(subset=['install_date', 'warranty_until', 'last_calibration_date', 'last_service_date'],
                   inplace=True)

    print("\nАналитика...")
    df_warranty = process_sync.filter_under_warranty_sync()
    df_problems = process_sync.problems_of_clinics_sync(process_sync.df)
    df_calibrations = process_sync.calibration_dates_sync(process_sync.df)
    df_pivot = process_sync.create_pivot_table_sync(process_sync.df)

    print("\nСохранение...")
    df_warranty.sort_values('warranty_until').to_excel('sync_1_filter_warranty.xlsx', index=False)
    df_problems.to_excel('sync_2_clinic_problems.xlsx', index=False)
    df_calibrations.sort_values('days_since_last_calibration', ascending=False).to_excel(
        'sync_3_calibration_report.xlsx', index=False)

    pivot_path = 'sync_4_pivot_table.xlsx'
    df_pivot.to_excel(pivot_path, index=False)
    _format_excel_pivot(pivot_path)

    total_time = time.time() - start_time
    print(f"\nСинхронная версия завершена за {total_time:.2f} секунд")
    return total_time


async def run_async(file_list):
    """Асинхронная обработка всех файлов.

    Args:
        file_list: Список путей к Excel файлам.

    Returns:
        Время выполнения в секундах.
    """

    print("ЗАПУСК АСИНХРОННОЙ ВЕРСИИ\n" + "=" * 60)
    start_time = time.time()

    process_async = MedicalDevices('combined')
    print("\nЧтение файлов (асинхронно)...")

    await process_async.read_all_files(file_list)

    process_async.parse_dates()
    process_async.status_normalization()
    process_async.checking_calibration_date()
    process_async.checking_service_date()
    process_async.df.dropna(subset=['install_date', 'warranty_until', 'last_calibration_date', 'last_service_date'],
                          inplace=True)

    print("\nАналитика (параллельно)...")
    task_start = time.time()
    df_warranty, df_problems, df_calibrations, df_pivot = await asyncio.gather(
        process_async.filter_under_warranty(),
        process_async.problems_of_clinics(process_async.df),
        process_async.calibration_dates(process_async.df),
        process_async.create_pivot_table(process_async.df)
    )
    print(f"Анализ выполнен за {time.time() - task_start:.2f} сек")

    print("\nАсинхронное сохранение...")
    save_start = time.time()

    await asyncio.gather(
        asyncio.to_thread(df_warranty.sort_values('warranty_until').to_excel, 'task1_filter_warranty.xlsx',
                          index=False),
        asyncio.to_thread(df_problems.to_excel, 'task2_clinic_problems.xlsx', index=False),
        asyncio.to_thread(df_calibrations.sort_values('days_since_last_calibration', ascending=False).to_excel,
                          'task3_calibration_report.xlsx', index=False),
        asyncio.to_thread(_save_and_format_pivot, df_pivot, 'task4_pivot_table.xlsx')
    )
    print(f"Сохранение выполнено за {time.time() - save_start:.2f} сек")

    total_time = time.time() - start_time
    print(f"\nАсинхронная версия завершена за {total_time:.2f} секунд")
    return total_time


def main():
    """Главная функция запуска сравнения синхронной и асинхронной версий."""
    file_list = [f'medical_diagnostic_devices_{i}.xlsx' for i in range(1, 11)]

    sync_time = run_sync(file_list)
    async_time = asyncio.run(run_async(file_list))

    print("\n" + "=" * 60)
    print("СРАВНЕНИЕ ВРЕМЕНИ ВЫПОЛНЕНИЯ")
    print("=" * 60)
    print(f"Синхронная версия: {sync_time:.2f} сек")
    print(f"Асинхронная версия: {async_time:.2f} сек")
    print(f"Разница: {abs(sync_time - async_time):.2f} сек")

    if async_time < sync_time:
        print(f"Асинхронная версия БЫСТРЕЕ на {((sync_time - async_time) / sync_time * 100):.1f}%")
    else:
        print(
            f"Синхронная версия быстрее (или равно). Разница: {((async_time - sync_time) / async_time * 100):.1f}%")
    print("=" * 60)


if __name__ == "__main__":
    main()
